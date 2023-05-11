import openpyxl
import statistics
import math
from scipy.stats import binom
from scipy.stats import zscore
from scipy.stats import norm

book = openpyxl.load_workbook("proprob.xlsx")
ALLsheet = []
sheet1 = book['ก.ค.65']
sheet2 = book['ส.ค.65']
sheet3 = book['ก.ย.65']
sheet4 = book['ต.ค.65']
dataAGEALL = []
dataAGE1 = []
dataAGE2 = []
dataAGE3 = []
dataAGE4 = []
dataALLtype =[]
dataA = [[],[],[],[]]
dataB = [[],[],[],[]]
dataC = [[],[],[],[]]
dataD = [[],[],[],[]]
dataE = [[],[],[],[]]
dataF = [[],[],[],[]]
dataG = [[],[],[],[]]

def main():
    crateSheet()
    crateDATAALL()
    crateALLType()
    for i in range(0,4):
        crateDATA(ALLsheet[i], dataAGEALL[i])
    for j in range(0,4):
        for k in range(0,7):
            crateTYPE(ALLsheet[j], j, dataALLtype[k], k)
    CallForUnit2_1_4(0,4)
    Unit2_5(dataAGEALL)
    Unit2_6()
    Unit3_1(dataAGEALL[2])
    Unit4_1(dataAGEALL[2])
    Unit5_1()
    Unit6_1()
    Unit7_1()
    Unit8_1()

def crateSheet():
    ALLsheet.append(sheet1)
    ALLsheet.append(sheet2)
    ALLsheet.append(sheet3) 
    ALLsheet.append(sheet4)

def crateDATAALL():
    dataAGEALL.append(dataAGE1)
    dataAGEALL.append(dataAGE2)
    dataAGEALL.append(dataAGE3)
    dataAGEALL.append(dataAGE4)

def crateALLType():
    dataALLtype.append(dataA)
    dataALLtype.append(dataB)
    dataALLtype.append(dataC)
    dataALLtype.append(dataD)
    dataALLtype.append(dataE)
    dataALLtype.append(dataF)
    dataALLtype.append(dataG)

def crateDATA(sheet, dataAGE):
    num_rows = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        age = row[4]
        if age is not None:
            dataAGE.append(age)
            num_rows += 1

def crateTYPE(sheet, j, dataType, k):
    num_rows = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        types = row[5]
        age = row[4]
        if types == 'A' and k == 0:
            dataType[j].append(age)
            num_rows += 1
        elif types == 'B' and k == 1:
            dataType[j].append(age)
            num_rows += 1
        elif types == 'C' and k == 2:
            dataType[j].append(age)
            num_rows += 1
        elif types == 'D' and k == 3:
            dataType[j].append(age)
            num_rows += 1
        elif types == 'E' and k == 4:
            dataType[j].append(age)
            num_rows += 1
        elif types == 'F' and k == 5:
            dataType[j].append(age)
            num_rows += 1
        elif types == 'G' and k == 6:
            dataType[j].append(age)
            num_rows += 1

def calcuAvg(dataAge, Sheetnow):
    Avgdev = 0
    for i in range(len(dataAge)):
        Avgdev += dataAge[i];
    return Avgdev/len(dataAge), Sheetnow

def calcuSta(dataAge, Sheetnow):
    stdev = statistics.stdev(dataAge) # หาค่าส่วนเบี่ยงเบนอายุผู้เสียชีวิต
    return stdev, Sheetnow

def calcuVar(dataAge, Sheetnow):
    variance = statistics.variance(dataAge)#หาค่าความแปรปรวนอายุผู้เสียชีวิต
    return variance, Sheetnow

def calcuMed(dataAge):
    median_value = statistics.median(dataAge)
    return median_value

def calcuMod(dataAge):
    mode_value = statistics.mode(dataAge)
    return mode_value

def CallForUnit2_1_4(S,E):
        for j in range(S,E):
            if j == 0:
                sheetsend = "กรกฏาคม:"
            elif j== 1:
                sheetsend = "สิงหาคม:"
            elif j== 2:
                sheetsend = "กันยายน:"
            elif j== 3:
                sheetsend = "ตุลาคม:"
            print("*"*70)
            print("Average age",calcuAvg(dataAGEALL[j], sheetsend)[1],calcuAvg(dataAGEALL[j], sheetsend)[0])
            print("Standard deviation",calcuSta(dataAGEALL[j], sheetsend)[1],calcuSta(dataAGEALL[j], sheetsend)[0])
            print("Variance",calcuVar(dataAGEALL[j], sheetsend)[1],calcuVar(dataAGEALL[j], sheetsend)[0])
            print("Median:",calcuMed(dataAGEALL[j]))
            print("Mode:",calcuMod(dataAGEALL[j]))

def Unit2_5(L1):
    LFTM = []
    Avgdev = 0
    for i in range(0,2):
        for j in range(len(L1[i])):
            LFTM.append(L1[i][j])
            Avgdev += L1[i][j]
    Avg = Avgdev/(len(L1[0]) + len(L1[1]))# หาค่าเฉลี่ยอายุผู้เสียชีวิต
    stdev = statistics.stdev(LFTM)# หาค่าส่วนเบี่ยงเบนอายุผู้เสียชีวิต
    variance = statistics.variance(LFTM)#หาค่าความแปรปรวนอายุผู้เสียชีวิต
    median_value = statistics.median(LFTM)
    mode_value = statistics.mode(LFTM)
    print("*"*70)
    print("Average age เดือนกรกฎาคมถึงเดือนสิงหาคม:", Avg)
    print("Standard deviation เดือนกรกฎาคมถึงเดือนสิงหาคม:", stdev)
    print("Variance เดือนกรกฎาคมถึงเดือนสิงหาคม:", variance)
    print("Median:", median_value)
    print("Mode:", mode_value)

def Unit2_6():
    AGE60 = 0
    ALLMEM = 0
    for i in range(0,4):
        ALLMEM += len((dataAGEALL[i]))
        for j in range(len(dataAGEALL[i])):
            if dataAGEALL[i][j] == 60:
                AGE60 += 1
    print("*"*70)
    print("ความน่าจะเป็นที่ตำรวจจะเสียชีวิตตอนอายุ60ปีตั้งแต่เดือนกรกฏาคมถึงเดือนตุลาคม", AGE60/ALLMEM)

def Unit3_1(dataAge):
    population = len(dataAge) #จำนวนคนทีตายในเดือนกันยายน
    dead = 0
    for i in dataAge:
        if i <60:
            dead += 1 #จำนวรคนืที่ตายก่อน 60ปี ในเดือนกันยายน
    prob0 = (math.comb(dead, 0)*math.comb(population-dead, 3))/math.comb(population, 3)
    prob1 = (math.comb(dead, 1)*math.comb(population-dead, 2))/math.comb(population, 3)
    prob2 = (math.comb(dead, 2)*math.comb(population-dead, 1))/math.comb(population, 3)
    prob3 = (math.comb(dead, 3)*math.comb(population-dead, 0))/math.comb(population, 3)
    print("*"*70)
    print("ความน่าจะเป็นของคนที่เสียชีวิตก่อนอายุ60ปีที่สุ่มมาเป็น 0 คน :", prob0) #ข้อย่อย ก
    print("ความน่าจะเป็นของคนที่เสียชีวิตก่อนอายุ60ปีที่สุ่มมาเป็น 1 คน :", prob1) #ข้อย่อย ก
    print("ความน่าจะเป็นของคนที่เสียชีวิตก่อนอายุ60ปีที่สุ่มมาเป็น 2 คน :", prob2) #ข้อย่อย ก
    print("ความน่าจะเป็นของคนที่เสียชีวิตก่อนอายุ60ปีที่สุ่มมาเป็น 3 คน :", prob3) #ข้อย่อย ก
    print("ความน่าจะเป็นของคนที่เสียชีวิตก่อนอายุ60ปีไม่น้อยกว่า 2 คน :", prob2+prob3) #ข้อย่อย ข

def Unit4_1(dataAge):
    population = len(dataAge)
    dead = 0
    for i in dataAge:
        if i < 80:
            dead+=1
    prob_dead = dead/population
    probability = binom.pmf(2, population, prob_dead)
    print("*"*70)
    print("ความน่าจะเป็น ", probability)

def Unit5_1():
    x = 80 #อายุของตำรวจที่เสียชีวิต
    u = calcuAvg(dataAGEALL[0], None)[0] #ตำรวจที่เสียชีวิตเดือนกรกฏาคมมีอายุเฉลี่ย
    variance = calcuVar(dataAGEALL[0], None)[0] #ความแปรปรวน
    z_score = (x - u) / math.sqrt(variance)
    probability = norm.cdf(z_score)
    print("*"*70)
    print("P( x > 80) = P( Z >",z_score,")")
    print("Probability:", 1-probability)

def Unit6_1():
    u = calcuAvg(dataAGEALL[3], None)[0]
    SD = math.sqrt(calcuVar(dataAGEALL[3], None)[0])
    n = 50
    x = 70
    z_score = (x - u) / (SD/math.sqrt(n))
    probability = norm.cdf(z_score)
    print("P( x < 70) = P( Z <",z_score,")")
    print("Probability:", probability)
    print("*"*70)

def Unit7_1():
    n = 40
    SD = math.sqrt(calcuVar(dataAGEALL[1], None)[0])
    x = calcuAvg(dataAGEALL[1], None)[0]
    z = 2.575
    lower = x - (z * SD/math.sqrt(n))
    upper = x + (z * SD/math.sqrt(n))
    print(lower , " < u < " , upper)
    print("*"*70)

def Unit8_1():
    u = calcuAvg(dataALLtype[2][2], None)[0] #อายุเฉลี่ยคนตายด้วยปอดเดือนกันยา
    x = 75
    SD = math.sqrt(calcuVar(dataALLtype[2][2], None)[0]) #ค่าเบี่ยงเบนคนตายด้วยปอดเดือนกันยา
    n = 36
    z_score = (x - u) / (SD/math.sqrt(n))
    print("z-score = ",z_score)
main()
